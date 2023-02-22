import json
import requests

from fastapi import FastAPI
from fastapi.responses import Response

from openpyxl import Workbook
from tempfile import NamedTemporaryFile


app = FastAPI()

@app.get('/create-excel')
def excel_creator(ticker:str,start_date:int,end_date:int):

        param = {'ticker':ticker,'start_date':start_date,'end_date':end_date}
        res = requests.get('http://0.0.0.0/get-ticker',param)

        financials = json.loads(res.json())


        wb = Workbook()
        sheet = wb.worksheets[0]

        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['P'].width = 25
        sheet.column_dimensions['Y'].width = 25

        row_index = 1

        for year in financials:

            statements = financials[year]

            sheet.cell(row=row_index,column=1 ,value=year)

            greater_length = 0

            for statement_name in statements:
                original_row_index = row_index

                if statement_name == 'Income':
                    column_index = 3
                elif statement_name == 'Balance':
                    column_index = 14
                elif statement_name == 'Cash':
                    column_index = 23

                statement = statements[statement_name]

                title = statement['title']

                sheet.cell(row=original_row_index+1,column=column_index,value=title)

                del statement['title']

                for concept_title in statement:

                    sheet.cell(row=original_row_index+3,column=column_index+1,value=concept_title)

                    concepts = statement[concept_title]

                    for concept in concepts:
                        value = concepts[concept]

                        sheet.cell(row=original_row_index+4,column=column_index+2,value=concept)
                        sheet.cell(row=original_row_index+4,column=column_index+3,value=value)
                        original_row_index+=1

                    original_row_index+=2

                difference_by_length = original_row_index-row_index
                if difference_by_length > greater_length:
                    greater_length = difference_by_length

            else:
                row_index+= greater_length+4


        with NamedTemporaryFile() as tmp:
                file_path = tmp.name
                wb.save(file_path)
                stream = tmp.read()
        
        return Response(content=stream,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
