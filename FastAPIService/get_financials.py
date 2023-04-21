import json
import os
import boto3
import re
import requests
import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook

from tempfile import NamedTemporaryFile

from fastapi import FastAPI,Response,HTTPException
from fastapi.exceptions import RequestValidationError
from fastapi.responses import Response as file_response,JSONResponse

from exception_handler_functions import error_validation_response,verify_statement_existence

app = FastAPI()


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    return JSONResponse(error_validation_response(exc),status_code=400,headers={'Access-Control-Allow-Origin':'*'})

current_year = datetime.date.today().year

@app.get("/get-ticker")
def get_ticker(ticker:str,start_date:int,end_date:int,response:Response):

    ticker = ticker.strip(' ')

    response.headers['Access-Control-Allow-Origin'] = '*'
    cors_policy_error_400 = {'Access-Control-Allow-Origin':'*'}


    if start_date > end_date or start_date < 2013:

        raise HTTPException(status_code=400,detail="The beginning year must be equal or less than the final year and greater than 2013",headers=cors_policy_error_400) 
    
    elif end_date > current_year:
        end_date = current_year+1

    sec_search_endpoint = 'https://efts.sec.gov/LATEST/search-index'

    header_req = {
    "user-Agent":os.getenv('USER_AGENT'),
    "accept-Encoding":"gzip, deflate, br",
    "origin":"https://www.sec.gov",
    "content-type":"application/x-www-form-urlencoded; charset=UTF-8",
    "cache-control": "no-cache",
    "referer":"https://www.sec.gov/"
    }


    statements_results = {}

    def search_cik(ticker):

            payload = json.dumps({"keysTyped":ticker})

            req_index = requests.post(url=sec_search_endpoint,data=payload,headers=header_req)

            try:
                first_result = req_index.json()["hits"]["hits"][0]
                if ticker in first_result['_source']['tickers']:

                    cik_without_0 = first_result['_id']
                    cik = '0'*(10-len(cik_without_0))+cik_without_0

                    return cik

                else:
                    raise HTTPException(status_code=400,detail="Provisioned ticker doesn't exist",headers=cors_policy_error_400) 

            except:
                raise HTTPException(status_code=400,detail="Provisioned ticker doesn't exist",headers=cors_policy_error_400) 

    year_order = {}
                
    def check_years(date_start,date_end):

        years_difference = date_end-date_start

        def years_to_order(differencial_year,year_order):

            year_order[differencial_year] = {}

            return differencial_year

        years_to_get = [ years_to_order(str(date_end-difference),year_order) for difference in range(years_difference+1)]

        if search_cik(ticker):

            table = boto3.resource('dynamodb').Table('fabri_app')

            response_item = table.get_item(
            Key={
                'ticker':ticker
                }
            )

            if 'Item' in response_item:
                year_to_check = []

                dates_attr = response_item['Item']

                attributes = {}

                for date in dates_attr:
                    year = date.split('-')[0]
                    if year in attributes:
                        attributes[year].append(date)
                    else:
                        attributes[year] = [date]

                for year in years_to_get:

                    if year not in attributes:
                
                        year_to_check.append(year)

                    else:

                        dates = attributes[year]

                        year_position = year_order[year]

                        for date in dates:

                            income = verify_statement_existence(dates_attr,date,'Income')
                            balance = verify_statement_existence(dates_attr,date,'Balance')
                            cash = verify_statement_existence(dates_attr,date,'Cash')

                            year_position[date] = {'Income':income,'Balance':balance,'Cash':cash}

                return year_to_check

            else:
                table.put_item(
                    Item={'ticker':ticker}
                )

                return years_to_get


    def filling_date_regular_expression(date):
        date_components = date.split('-')

        dict_months = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
        regular_expression_date = dict_months[int(date_components[1])]+f'.*{date_components[2]}.*{date_components[0]}'

        return {'regrex_expression':regular_expression_date}


    def get_10_k_links(ticker,start_date,end_date,valid_list):

        cik = search_cik(ticker)

        payload = json.dumps({"entityName":cik,"filter_forms":"10-K","startdt":f"{str(start_date)}-01-01","enddt":f"{str(end_date+1)}-12-31"})

        req = requests.post(sec_search_endpoint,headers=header_req,data=payload)

        list_fillings = req.json()['hits']['hits']

        list_adsh = [{'date':filling_date_regular_expression(filling['_source']['period_ending']),'url':f'https://www.sec.gov/Archives/edgar/data/{cik}/'+filling['_source']['adsh'].replace('-','')} for filling in list_fillings if filling['_source']['file_type'] == '10-K' and int(filling['_source']['period_ending'].split('-')[0]) <= end_date and int(filling['_source']['period_ending'].split('-')[0]) >= start_date ]


        if list_adsh:
            last_fillings_year = int(list_adsh[0]['date']['regrex_expression'].split('.')[2][1:])

            for year in valid_list.copy():
                if int(year) > last_fillings_year:
                    valid_list.remove(year)

            return (list_adsh)

        else:
            raise HTTPException(status_code=400,detail="Ticker doesn't represent a US company",headers=cors_policy_error_400)


    new_financial_data = {}

    regex_date = re.compile(r"[A-Za-z][A-Za-z][A-Za-z]\. \d\d, \d\d\d\d")
    regex_year = re.compile(r"\d\d\d\d")
    regex_day = re.compile(r"\d\d,")
    regex_month = re.compile(r"[A-Za-z][A-Za-z][A-Za-z]", re.IGNORECASE)

    regex_cell_number_class = re.compile('nump|num|text',re.I)
    regex_first_income_row_class = re.compile('nump|num',re.I)

    number_month_dict = {'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04', 'may': '05', 'jun': '06',
                        'jul': '07', 'aug': '08', 'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'}


    def valid_format_date(input_date):
        date = regex_date.findall(input_date)[0]
        year = regex_year.findall(date)[0]
        day = regex_day.findall(date)[0][0:2]
        month = number_month_dict[regex_month.findall(date)[0].lower()]
        date_result = f'{year}-{month}-{day}'
        return date_result,year



    def parse_statement(type,non_balance,file_url,date,end_date,valid_years_list,is_income = False):

        req = requests.get(file_url,headers=header_req).text

        content = BeautifulSoup(req,features="html.parser")
        rows_list = content.find('table',{"class": "report"}).find_all('tr')

        currency_desc = rows_list[0].find('th',attrs={'class':'tl'}).getText()

        if non_balance:
            row_years_index = 1
            rows__first_index_concepts = 2
            if is_income and not rows_list[2].find_all('td',attrs={'class':regex_first_income_row_class}) :
                rows__first_index_concepts = 3

        if not non_balance:
            row_years_index = 0
            rows__first_index_concepts = 1


        years_row = rows_list[row_years_index].find_all('th',attrs={'class':'th'})

        
        years = [ regex_date.findall(year_row.getText())[0] for year_row in years_row if regex_date.findall(year_row.getText())]

        for index,year in enumerate(reversed(years)):
            if re.search(f'{date}',year,re.I):

                filling_year = len(years)-(index+1)

                valid_years = []
                invalid_years_index = []

                years_row_to_use = years[filling_year:]

                for index_year in range(0,len(years_row_to_use)):

                    current_date = years_row_to_use[index_year]

                    date_formated = valid_format_date(current_date)
                    valid_date = date_formated[0]
                    current_year = date_formated[1]


                    if current_year not in valid_years_list or int(current_year) < end_date :
                        invalid_years_index.insert(0,index_year)
                        continue

                    if valid_date in year_order[current_year]:
                        if type in year_order[current_year][valid_date]:
                            invalid_years_index.insert(0,index_year)
                            continue
                        else:
                            year_order[current_year][valid_date][type] = {'title':currency_desc}
                            new_financial_data[valid_date][type] = {'title':currency_desc}

                    else:
                        year_order[current_year][valid_date] = {type:{'title':currency_desc}}
                        new_financial_data[valid_date] = {type:{'title':currency_desc}}


                    valid_years.append((valid_date,current_year))


                break   


        concept_title = 'FirstBlock'
        concept_titles_used = set({})

        for row in rows_list[rows__first_index_concepts:]:

            cells_total = row.find_all('td',attrs={'class':regex_cell_number_class})

            len_cells = len(years)-len(cells_total)

            cells = cells_total[filling_year-len_cells:]
            
            concept = row.find('td',attrs={'class':'pl'})

            is_concept_title = concept.find('strong') if concept else True

            if not is_concept_title and cells and concept_title:

                    for invalid_index in invalid_years_index:
                            cells.pop(invalid_index)

                    concept_text = concept.getText()

                    for index,cell in enumerate(cells):
                        valid_index = valid_years[index]

                        valid_date = valid_index[0]

                        valid_year = valid_index[1]
   
                        cell_class = cell.get('class')[0]

                        non_digit_filter = re.sub('\D','',cell.getText())

                        if cell_class == 'num':
                            non_digit_filter = f'({non_digit_filter})'

                        try:

                            new_financial_data[valid_date][type][concept_title][concept_text] = non_digit_filter
                            year_order[valid_year][valid_date][type][concept_title][concept_text] = non_digit_filter

                        except:
                            new_financial_data[valid_date][type][concept_title] = {concept_text: non_digit_filter}
                            year_order[valid_year][valid_date][type][concept_title] = {concept_text: non_digit_filter}


            else:
                if is_concept_title:
                    try:
                        concept_title = concept.getText()

                        if concept_title in concept_titles_used:
                            concept_title = None
                        else:
                            concept_titles_used.add(concept_title)

                    except:
                        continue        


    
    def get_statements(ticker,start_date,end_date,valid_years_list):

        fillings = get_10_k_links(ticker,start_date,end_date,valid_years_list)

        if not valid_years_list:
            return False

        for filling_dict in fillings:
            filling = filling_dict['url']
            filling_date = filling_dict['date']['regrex_expression']

            filling_summary = filling+'/FilingSummary.xml'

            req = requests.get(filling_summary,headers=header_req)

            reports = BeautifulSoup(req.text,features='xml').find_all('Report')[1:10]
            income_state = False
            balance_state = False
            flow_state = False

            for report in reports:
                report_name = report.find('ShortName').getText().upper()
                html_file_name = report.find('HtmlFileName').getText()
                file_url = filling+f'/{html_file_name}'
                if "OPERATION" in report_name or "INCOME" in report_name or "EARNING" in report_name:
                    if not income_state:
                        parse_statement('Income',True,file_url,filling_date,start_date,valid_years_list,True)
                        income_state = True

                if "BALANCE" in report_name and "SHEET" in report_name:
                    if not balance_state:

                        parse_statement('Balance',False,file_url,filling_date,start_date,valid_years_list)
                        balance_state = True

                if "CASH FLOW" in report_name:
                    if not flow_state:
                        parse_statement('Cash',True,file_url,filling_date,start_date,valid_years_list)
                        flow_state = True


        return new_financial_data


    new_financial_result = get_statements(ticker,start_date,end_date,check_years(start_date,end_date))
    

    for year in year_order:
        dates = year_order[year]
        for date in dates:
            statements_results[date] = dates[date]


    if new_financial_result:
        sqs = boto3.client('sqs')

        message = json.dumps(new_financial_data)

        sqs_attribute = {
                'ticker': {
                    'DataType': 'String',
                    'StringValue': ticker
                }
        }

        sqs.send_message(
            QueueUrl='https://sqs.us-west-2.amazonaws.com/870828436064/TickerQueue',
            MessageBody=message,
            MessageAttributes=sqs_attribute
        )

  
    return json.dumps(statements_results)



@app.get('/create-excel')
def excel_creator(ticker:str,start_date:int,end_date:int):

        param = {'ticker':ticker,'start_date':start_date,'end_date':end_date}
        res = requests.get('http://127.0.0.1:8000/get-ticker',param)

        financials = json.loads(res.json())

        wb = Workbook()
        sheet = wb.worksheets[0]

        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['P'].width = 25
        sheet.column_dimensions['Y'].width = 25

        row_index = 1

        for year in financials:

            statements = financials[year]

            sheet.cell(row=row_index,column=1 ,value=year)

            greater_length = 0

            for statement_name in statements:
                original_row_index = row_index

                if statement_name == 'Income':
                    column_index = 3
                elif statement_name == 'Balance':
                    column_index = 14
                elif statement_name == 'Cash':
                    column_index = 23

                statement = statements[statement_name]

                title = statement['title']

                sheet.cell(row=original_row_index+1,column=column_index,value=title)

                del statement['title']

                for concept_title in statement:

                    sheet.cell(row=original_row_index+3,column=column_index+1,value=concept_title)

                    concepts = statement[concept_title]

                    for concept in concepts:
                        value = concepts[concept]

                        sheet.cell(row=original_row_index+4,column=column_index+2,value=concept)
                        sheet.cell(row=original_row_index+4,column=column_index+3,value=value)
                        original_row_index+=1

                    original_row_index+=2

                difference_by_length = original_row_index-row_index
                if difference_by_length > greater_length:
                    greater_length = difference_by_length

            else:
                row_index+= greater_length+4


        with NamedTemporaryFile() as tmp:
                file_path = tmp.name
                wb.save(file_path)
                stream = tmp.read()
        
        return file_response(content=stream,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
